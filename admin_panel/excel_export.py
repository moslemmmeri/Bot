# admin_panel/excel_export.py
# تولید فایل‌های Excel از داده‌های سفارشات با استفاده از openpyxl
# شامل: فرمت‌بندی سلول‌ها، چندین شیت، رنگ‌بندی و آمار

import os
import io
import tempfile
import traceback  # ✅ اضافه شد برای traceback کامل
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from logger_config import logger
from utils.error_handler import log_general_error  # ✅ اضافه شد


# ============================================================
# تنظیمات
# ============================================================

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')


def ensure_export_dir():
    """ایجاد پوشه خروجی در صورت عدم وجود"""
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
    return EXPORT_DIR


# ============================================================
# استایل‌ها
# ============================================================

# رنگ‌ها
COLOR_HEADER = "1F4E79"      # آبی تیره
COLOR_HEADER_TEXT = "FFFFFF" # سفید
COLOR_ROW_ODD = "D6E4F0"     # آبی روشن
COLOR_ROW_EVEN = "FFFFFF"    # سفید
COLOR_BORDER = "4472C4"      # آبی متوسط
COLOR_PENDING = "ED7D31"     # نارنجی
COLOR_PAID = "70AD47"        # سبز
COLOR_COMPLETED = "2E75B6"   # آبی
COLOR_TOTAL = "4472C4"       # آبی

# فونت‌ها
FONT_HEADER = Font(name='B Nazanin', size=12, bold=True, color=COLOR_HEADER_TEXT)
FONT_NORMAL = Font(name='B Nazanin', size=10)
FONT_BOLD = Font(name='B Nazanin', size=10, bold=True)
FONT_TITLE = Font(name='B Nazanin', size=14, bold=True)

# Alignment
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

# Border
BORDER_THIN = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)


# ============================================================
# توابع اصلی
# ============================================================

def create_orders_excel(
    orders: List[Dict],
    title: str = "گزارش سفارشات",
    include_stats: bool = True,
    include_chart: bool = True
) -> Optional[str]:
    """
    ایجاد فایل Excel از لیست سفارشات با چندین شیت
    
    پارامترها:
        orders: لیست دیکشنری‌های سفارشات
        title: عنوان فایل
        include_stats: آیا شیت آمار اضافه شود
        include_chart: آیا نمودار اضافه شود
    
    بازگشت: مسیر فایل ایجادشده یا None در صورت خطا
    """
    if not orders:
        logger.warning("No orders provided for Excel export")
        return None
    
    try:
        ensure_export_dir()
        
        # ایجاد Workbook
        wb = Workbook()
        
        # حذف شیت پیش‌فرض
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # ۱. شیت اصلی سفارشات
        orders_sheet = wb.create_sheet("سفارشات")
        _populate_orders_sheet(orders_sheet, orders)
        
        # ۲. شیت آمار (اختیاری)
        if include_stats:
            stats_sheet = wb.create_sheet("آمار و خلاصه")
            _populate_stats_sheet(stats_sheet, orders)
        
        # ۳. شیت نمودار (اختیاری)
        if include_chart and len(orders) > 1:
            chart_sheet = wb.create_sheet("نمودار")
            _populate_chart_sheet(chart_sheet, orders)
        
        # ذخیره فایل
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"orders_report_{timestamp}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)
        
        wb.save(filepath)
        logger.info(f"Excel file created: {filepath}")
        
        return filepath
        
    except Exception as e:
        log_general_error(  # ✅ استفاده از log_general_error با traceback کامل
            f"Error creating Excel file: {str(e)}",
            traceback=traceback.format_exc()
        )
        return None


def _populate_orders_sheet(sheet, orders: List[Dict]):
    """پر کردن شیت سفارشات"""
    
    # ستون‌ها
    columns = [
        ('شناسه', 12),
        ('شناسه کاربر', 15),
        ('نام کاربر', 25),
        ('سرویس', 30),
        ('مبلغ (ریال)', 18),
        ('وضعیت', 15),
        ('کد رهگیری', 20),
        ('تاریخ ثبت', 20),
        ('یادداشت ادمین', 30),
    ]
    
    # تنظیم عرض ستون‌ها
    for col_idx, (col_name, width) in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = width
        # تنظیم هدر
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    
    # پر کردن داده‌ها
    for row_idx, order in enumerate(orders, 2):
        # رنگ زمینه بر اساس ردیف
        fill_color = COLOR_ROW_ODD if row_idx % 2 == 0 else COLOR_ROW_EVEN
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # استخراج داده‌ها
        order_id = order.get('id', '')
        user_id = order.get('user_id', '')
        fullname = _get_fullname(order)
        service_name = _get_service_name(order)
        amount = order.get('payment_amount', 0) or 0
        status = order.get('status', 'pending')
        tracking = order.get('tracking_code', '')
        created = order.get('created_at', '')
        admin_note = order.get('admin_note', '')
        
        # وضعیت فارسی
        status_persian = _get_status_persian(status)
        
        # مقداردهی سلول‌ها
        row_data = [order_id, user_id, fullname, service_name, amount, status_persian, tracking, created, admin_note]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = FONT_NORMAL
            cell.fill = fill
            cell.alignment = ALIGN_CENTER if col_idx in [1, 2, 5, 6] else ALIGN_RIGHT if col_idx == 5 else ALIGN_LEFT
            cell.border = BORDER_THIN
    
    # فریز کردن ردیف اول
    sheet.freeze_panes = 'A2'
    
    # اضافه کردن فیلتر
    sheet.auto_filter.ref = sheet.dimensions


def _populate_stats_sheet(sheet, orders: List[Dict]):
    """پر کردن شیت آمار"""
    
    # عنوان
    sheet.merge_cells('A1:D1')
    cell = sheet['A1']
    cell.value = "📊 خلاصه آمار سفارشات"
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER
    
    # آمار کلی
    total_orders = len(orders)
    total_amount = sum(o.get('payment_amount', 0) or 0 for o in orders)
    avg_amount = total_amount / total_orders if total_orders > 0 else 0
    paid_orders = len([o for o in orders if o.get('status') in ['paid', 'completed']])
    pending_orders = len([o for o in orders if o.get('status') == 'pending'])
    
    stats = [
        ('📦 تعداد کل سفارشات', total_orders),
        ('💰 مجموع مبلغ', f"{total_amount:,} ریال"),
        ('📊 میانگین مبلغ هر سفارش', f"{avg_amount:,.0f} ریال"),
        ('✅ سفارشات پرداخت‌شده', paid_orders),
        ('⏳ سفارشات در انتظار', pending_orders),
    ]
    
    row = 3
    for label, value in stats:
        sheet.cell(row=row, column=1).value = label
        sheet.cell(row=row, column=1).font = FONT_BOLD
        sheet.cell(row=row, column=2).value = value
        sheet.cell(row=row, column=2).font = FONT_NORMAL
        row += 1
    
    # تنظیم عرض ستون‌ها
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 25
    
    # آمار وضعیت‌ها (جدول دوم)
    row += 2
    sheet.cell(row=row, column=1).value = "📌 تفکیک وضعیت سفارشات"
    sheet.cell(row=row, column=1).font = FONT_BOLD
    row += 1
    
    # محاسبه آمار وضعیت‌ها
    status_counts = {}
    for order in orders:
        status = order.get('status', 'unknown')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    # جدول وضعیت‌ها
    sheet.cell(row=row, column=1).value = "وضعیت"
    sheet.cell(row=row, column=2).value = "تعداد"
    sheet.cell(row=row, column=1).font = FONT_HEADER
    sheet.cell(row=row, column=2).font = FONT_HEADER
    sheet.cell(row=row, column=1).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    sheet.cell(row=row, column=2).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    sheet.cell(row=row, column=1).alignment = ALIGN_CENTER
    sheet.cell(row=row, column=2).alignment = ALIGN_CENTER
    row += 1
    
    color_map = {
        'pending': COLOR_PENDING,
        'paid': COLOR_PAID,
        'completed': COLOR_COMPLETED,
    }
    
    for status, count in status_counts.items():
        status_persian = _get_status_persian(status)
        sheet.cell(row=row, column=1).value = status_persian
        sheet.cell(row=row, column=2).value = count
        sheet.cell(row=row, column=1).font = FONT_NORMAL
        sheet.cell(row=row, column=2).font = FONT_NORMAL
        
        color = color_map.get(status, COLOR_HEADER)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet.cell(row=row, column=1).fill = fill
        sheet.cell(row=row, column=2).fill = fill
        
        row += 1
    
    # سرویس‌های برتر (جدول سوم)
    row += 2
    sheet.cell(row=row, column=1).value = "🏆 سرویس‌های برتر"
    sheet.cell(row=row, column=1).font = FONT_BOLD
    row += 1
    
    # محاسبه تعداد سفارشات هر سرویس
    service_counts = {}
    for order in orders:
        service_name = _get_service_name(order)
        service_counts[service_name] = service_counts.get(service_name, 0) + 1
    
    # مرتب‌سازی و نمایش
    sorted_services = sorted(service_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    
    sheet.cell(row=row, column=1).value = "سرویس"
    sheet.cell(row=row, column=2).value = "تعداد سفارش"
    sheet.cell(row=row, column=1).font = FONT_HEADER
    sheet.cell(row=row, column=2).font = FONT_HEADER
    sheet.cell(row=row, column=1).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    sheet.cell(row=row, column=2).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    sheet.cell(row=row, column=1).alignment = ALIGN_CENTER
    sheet.cell(row=row, column=2).alignment = ALIGN_CENTER
    row += 1
    
    for service_name, count in sorted_services:
        sheet.cell(row=row, column=1).value = service_name
        sheet.cell(row=row, column=2).value = count
        sheet.cell(row=row, column=1).font = FONT_NORMAL
        sheet.cell(row=row, column=2).font = FONT_NORMAL
        row += 1


def _populate_chart_sheet(sheet, orders: List[Dict]):
    """پر کردن شیت نمودار"""
    
    # عنوان
    sheet.merge_cells('A1:D1')
    cell = sheet['A1']
    cell.value = "📊 نمودار سفارشات به تفکیک سرویس"
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER
    
    # محاسبه داده‌ها
    service_counts = {}
    for order in orders:
        service_name = _get_service_name(order)
        service_counts[service_name] = service_counts.get(service_name, 0) + 1
    
    sorted_services = sorted(service_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    
    if not sorted_services:
        sheet.cell(row=3, column=1).value = "هیچ داده‌ای برای نمایش وجود ندارد."
        return
    
    # نوشتن داده‌ها
    row = 3
    sheet.cell(row=row, column=1).value = "سرویس"
    sheet.cell(row=row, column=2).value = "تعداد سفارش"
    sheet.cell(row=row, column=1).font = FONT_HEADER
    sheet.cell(row=row, column=2).font = FONT_HEADER
    sheet.cell(row=row, column=1).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    sheet.cell(row=row, column=2).fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    sheet.cell(row=row, column=1).alignment = ALIGN_CENTER
    sheet.cell(row=row, column=2).alignment = ALIGN_CENTER
    row += 1
    
    for service_name, count in sorted_services:
        sheet.cell(row=row, column=1).value = service_name
        sheet.cell(row=row, column=2).value = count
        row += 1
    
    # ایجاد نمودار میل‌ای
    chart = BarChart()
    chart.type = "col"
    chart.title = "تعداد سفارشات به تفکیک سرویس"
    chart.y_axis.title = "تعداد سفارش"
    chart.x_axis.title = "سرویس"
    
    data = Reference(sheet, min_col=2, min_row=3, max_row=row-1, max_col=2)
    categories = Reference(sheet, min_col=1, min_row=4, max_row=row-1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    # تنظیم اندازه نمودار
    chart.width = 20
    chart.height = 12
    
    # افزودن نمودار به شیت
    sheet.add_chart(chart, "E3")


# ============================================================
# توابع کمکی
# ============================================================

def _get_fullname(order: Dict) -> str:
    """استخراج نام کامل کاربر از سفارش"""
    import json
    order_data = order.get('order_data', {})
    if isinstance(order_data, str):
        try:
            order_data = json.loads(order_data)
        except:
            order_data = {}
    
    fullname = order_data.get('fullname')
    if fullname:
        return fullname
    
    answers = order_data.get('answers', {})
    if answers:
        first_answer = next(iter(answers.values()), 'کاربر ناشناس')
        return first_answer
    
    return 'کاربر ناشناس'


def _get_service_name(order: Dict) -> str:
    """دریافت نام سرویس از سفارش"""
    from database import get_button_by_id
    
    button_id = order.get('button_id')
    if not button_id:
        return 'نامشخص'
    
    btn = get_button_by_id(button_id)
    if not btn:
        return f"سرویس {button_id}"
    
    if btn.get('parent_button_id'):
        parent = get_button_by_id(btn['parent_button_id'])
        if parent:
            return f"{parent['name']} > {btn['name']}"
    
    return btn['name']


def _get_status_persian(status: str) -> str:
    """تبدیل وضعیت به فارسی"""
    status_map = {
        'pending': '⏳ در انتظار پرداخت',
        'paid': '✅ پرداخت شده',
        'completed': '✅ تکمیل شده',
        'cancelled': '❌ لغو شده',
    }
    return status_map.get(status, status)


# ============================================================
# کلاس ExcelExporter برای مدیریت یکپارچه
# ============================================================

class ExcelExporter:
    """مدیریت یکپارچه خروجی Excel"""
    
    def __init__(self):
        ensure_export_dir()
    
    def export_orders(self, orders: List[Dict], title: str = "گزارش سفارشات") -> Optional[str]:
        """خروجی سفارشات به Excel"""
        return create_orders_excel(orders, title, include_stats=True, include_chart=True)
    
    def export_orders_simple(self, orders: List[Dict], title: str = "گزارش سفارشات") -> Optional[str]:
        """خروجی ساده سفارشات (بدون آمار و نمودار)"""
        return create_orders_excel(orders, title, include_stats=False, include_chart=False)
    
    def export_orders_with_stats(self, orders: List[Dict], title: str = "گزارش کامل سفارشات") -> Optional[str]:
        """خروجی کامل با آمار و نمودار"""
        return create_orders_excel(orders, title, include_stats=True, include_chart=True)
    
    def export_filtered_orders(self, orders: List[Dict], filter_name: str = "") -> Optional[str]:
        """خروجی سفارشات فیلترشده"""
        title = f"گزارش سفارشات {filter_name}".strip()
        return self.export_orders(orders, title)
    
    def cleanup_old_exports(self, days: int = 7) -> int:
        """حذف فایل‌های Excel قدیمی"""
        deleted = 0
        cutoff = datetime.now().timestamp() - (days * 24 * 3600)
        
        for filename in os.listdir(EXPORT_DIR):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(EXPORT_DIR, filename)
                if os.path.getmtime(filepath) < cutoff:
                    try:
                        os.remove(filepath)
                        deleted += 1
                    except Exception as e:
                        log_general_error(  # ✅ استفاده از log_general_error با traceback کامل
                            f"Error deleting old export {filename}: {str(e)}",
                            traceback=traceback.format_exc()
                        )
        
        return deleted


# ============================================================
# توابع راحت‌تر برای استفاده در پنل مدیریت
# ============================================================

def export_orders_excel(orders: List[Dict]) -> Optional[str]:
    """تابع راحت برای خروجی Excel سفارشات"""
    return create_orders_excel(orders, "گزارش سفارشات", include_stats=True, include_chart=True)


def export_orders_simple_excel(orders: List[Dict]) -> Optional[str]:
    """خروجی ساده Excel"""
    return create_orders_excel(orders, "گزارش سفارشات", include_stats=False, include_chart=False)


def get_excel_exporter() -> ExcelExporter:
    """دریافت آبجکت ExcelExporter"""
    return ExcelExporter()


__all__ = [
    'create_orders_excel',
    'export_orders_excel',
    'export_orders_simple_excel',
    'ExcelExporter',
    'get_excel_exporter',
    'ensure_export_dir',
]